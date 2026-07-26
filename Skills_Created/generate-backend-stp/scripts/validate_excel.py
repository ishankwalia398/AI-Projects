#!/usr/bin/env python
"""
Validates an Excel workbook: opens it, optionally checks expected sheet names exist,
re-saves the file (preserving existing cached values), and reports results as JSON.
Note: openpyxl does not recalculate formulas — it preserves whatever cached results exist.

Usage:
  python validate_excel.py <path> [expected_sheet1 expected_sheet2 ...]

Exit codes: 0 = success, 1 = validation errors or exception.
"""
import sys
import json


def validate(path, expected_sheets=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(json.dumps({
            "status": "errors_found",
            "errors": ["openpyxl not installed. Run: pip install openpyxl -q"]
        }))
        sys.exit(1)

    try:
        wb = load_workbook(path)
        errors = []

        if expected_sheets:
            for name in expected_sheets:
                if name not in wb.sheetnames:
                    errors.append(f"Missing sheet: '{name}'")

        for ws in wb.worksheets:
            if ws.max_row is None or ws.max_row <= 1:
                errors.append(f"Sheet '{ws.title}' appears to have no data rows")

        if errors:
            print(json.dumps({"status": "errors_found", "errors": errors}))
            sys.exit(1)
        else:
            wb.save(path)
            print(json.dumps({
                "status": "success",
                "sheets": wb.sheetnames,
                "path": str(path)
            }))
    except Exception as e:
        print(json.dumps({"status": "errors_found", "errors": [str(e)]}))
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({
            "status": "errors_found",
            "errors": ["Usage: python validate_excel.py <path> [sheet1 sheet2 ...]"]
        }))
        sys.exit(1)
    validate(sys.argv[1], sys.argv[2:] if len(sys.argv) > 2 else None)
