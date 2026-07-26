#!/usr/bin/env python3
"""
Automation Report Analysis Script
Analyzes test execution reports and generates RCA Excel output
"""

import json
import os
import sys
import zipfile
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


class TestFailure:
    """Represents a single test failure with all relevant details"""
    def __init__(self, test_name: str, test_dir: str):
        self.test_name = test_name
        self.test_dir = test_dir
        self.api_call = ""
        self.api_url = ""
        self.api_method = ""
        self.request_params = ""
        self.response_headers = {}
        self.response_body = ""
        self.failure_message = ""
        self.failure_reason = ""
        self.failure_category = ""
        self.solution = ""


class ReportAnalyzer:
    """Main analyzer class for test reports"""

    def __init__(self, zip_path: str, output_dir: str = None):
        self.zip_path = zip_path
        # Default to ~/.claude/outputs/ if no output_dir specified
        if output_dir is None:
            home_dir = Path.home()
            output_dir = home_dir / ".claude" / "outputs"
            output_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir = output_dir
        self.extract_dir = None
        self.failures: List[TestFailure] = []
        self.total_tests = 0
        self.total_passed = 0
        self.total_failed = 0

    def extract_report(self) -> Path:
        """Extract the zip file to a temporary directory"""
        # Use system temp directory for extraction to avoid cluttering outputs
        import tempfile
        temp_base = Path(tempfile.gettempdir())
        extract_path = temp_base / f"claude_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        extract_path.mkdir(parents=True, exist_ok=True)

        print(f"Extracting report to {extract_path}...")
        with zipfile.ZipFile(self.zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_path)

        self.extract_dir = extract_path
        return extract_path

    def find_test_directories(self) -> List[Path]:
        """Find all test directories in the extracted report"""
        tests_dir = self.extract_dir / "tests"
        if not tests_dir.exists():
            raise FileNotFoundError(f"Tests directory not found at {tests_dir}")

        test_dirs = [d for d in tests_dir.iterdir() if d.is_dir()]
        print(f"Found {len(test_dirs)} test directories")
        return test_dirs

    def parse_test_js(self, test_js_path: Path) -> Dict:
        """Parse the test.js file to extract JSON data"""
        with open(test_js_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Remove JavaScript variable declaration
        content = content.replace('var test = ', '').rstrip().rstrip(';')

        try:
            return json.loads(content)
        except json.JSONDecodeError as e:
            print(f"Error parsing {test_js_path}: {e}")
            return {}

    def check_503_or_upstream_failure(self, failure_message: str, test_data: Dict) -> bool:
        """Check if this failure is 503 or no healthy upstream related"""
        skip_patterns = [
            '503',
            'no healthy upstream',
            'NO HEALTHY UPSTREAM',
            'No healthy upstream'
        ]

        failure_text = failure_message.lower()
        for pattern in skip_patterns:
            if pattern.lower() in failure_text:
                return True

        # Also check in surrounding elements for 503 responses
        if 'reportElements' in test_data:
            for elem in test_data['reportElements']:
                title = str(elem.get('title', '')).lower()
                if '503' in title or 'no healthy upstream' in title:
                    return True

        return False

    def extract_api_details(self, test_data: Dict, failed_idx: int, test_dir: Path) -> Tuple[str, str, str, str, Dict, str]:
        """Extract API call, URL, method, params, response headers and body"""
        api_call = ""
        api_url = ""
        api_method = "Unknown"
        request_params = ""
        response_headers = {}
        response_body = ""

        # Search backwards from failure to find API call
        elements = test_data.get('reportElements', [])
        for i in range(failed_idx, max(0, failed_idx - 20), -1):
            title = elements[i].get('title', '') or ''
            if 'API Call:' in title:
                api_call = title
                # Extract endpoint from title
                parts = title.split('API Call:')
                if len(parts) > 1:
                    api_url = parts[1].strip()

                # Try to determine method from context
                if 'POST' in title.upper():
                    api_method = 'POST'
                elif 'GET' in title.upper():
                    api_method = 'GET'
                elif 'PUT' in title.upper():
                    api_method = 'PUT'
                elif 'DELETE' in title.upper():
                    api_method = 'DELETE'

                break

        # Look for response files
        response_files = list(test_dir.glob('_response_*.txt'))
        if response_files:
            # Use the first response file found
            response_file = response_files[0]
            try:
                with open(response_file, 'r', encoding='utf-8') as f:
                    response_data = json.load(f)

                # Extract mock response details
                if isinstance(response_data, list) and len(response_data) > 0:
                    # Find relevant mock based on API call
                    for mock in response_data:
                        stub_name = mock.get('stubName', '')
                        if api_url and stub_name in api_url:
                            response_headers = mock.get('responseHeaders', {})
                            response_body = json.dumps(mock.get('responseBody', {}), indent=2)
                            api_method = mock.get('httpMethod', api_method)
                            break

                    # If no match, use first mock
                    if not response_body:
                        mock = response_data[0]
                        response_headers = mock.get('responseHeaders', {})
                        response_body = json.dumps(mock.get('responseBody', {}), indent=2)
                        api_method = mock.get('httpMethod', api_method)

            except (json.JSONDecodeError, IOError) as e:
                print(f"Error reading response file {response_file}: {e}")

        # If no response file, try to extract from test elements
        if not response_body:
            for i in range(failed_idx - 1, max(0, failed_idx - 10), -1):
                elem_title = elements[i].get('title') or ''
                if 'Response' in elem_title or 'response' in elem_title:
                    response_body += f"{elem_title}\n"

        return api_call, api_url, api_method, request_params, response_headers, response_body

    def analyze_failure(self, failure: TestFailure, test_data: Dict = None):
        """Perform RCA on the failure and populate failure_reason, category, and solution"""
        failure_msg = failure.failure_message.lower()

        # Check for 503 or no healthy upstream first
        is_503_failure = test_data and self.check_503_or_upstream_failure(failure.failure_message, test_data)

        if is_503_failure or '503' in failure_msg or 'no healthy upstream' in failure_msg:
            failure.failure_category = "Service Unavailable 503"
            failure.failure_reason = "Service unavailable (503). The upstream service is not responding or has no healthy instances."
            failure.solution = "1. Check if the upstream service is running\n2. Verify load balancer health checks\n3. Check service deployment status\n4. Review service logs for crashes or errors\n5. Verify network connectivity to upstream service\n6. Check if service is overloaded or throttling requests"

        elif 'totalcount' in failure_msg and '0' in failure_msg:
            failure.failure_category = "Data Validation Failure"
            failure.failure_reason = "API returned empty result set. TotalCount is 0 when test expected at least 1 record."
            failure.solution = "1. Verify test data setup - ensure required records exist in the database\n2. Check if data cleanup ran before this test\n3. Verify API query parameters are correct\n4. Check database connection and query execution logs"

        elif 'field' in failure_msg and 'value is' in failure_msg:
            failure.failure_category = "Field Validation Failure"
            failure.failure_reason = "Response field validation failed. The actual value doesn't match expected value."
            failure.solution = "1. Review API implementation for this field\n2. Check if API contract changed\n3. Verify test expectations are still valid\n4. Update test assertion if API behavior is correct"

        elif 'not found' in failure_msg or '404' in failure_msg:
            failure.failure_category = "API Error 404"
            failure.failure_reason = "Resource not found (404). API endpoint or resource ID may be incorrect."
            failure.solution = "1. Verify the resource ID exists in the test environment\n2. Check API endpoint URL is correct\n3. Ensure test data setup completed successfully\n4. Review API routing configuration"

        elif '401' in failure_msg or 'unauthorized' in failure_msg:
            failure.failure_category = "Authentication Failure"
            failure.failure_reason = "Authentication failure. Token, KS, or credentials are invalid or expired."
            failure.solution = "1. Refresh authentication token before this test\n2. Verify user credentials are correct\n3. Check if session expired\n4. Review authentication service logs"

        elif '403' in failure_msg or 'forbidden' in failure_msg:
            failure.failure_category = "Authorization Failure"
            failure.failure_reason = "Authorization failure. User lacks required permissions for this operation."
            failure.solution = "1. Verify test user has correct roles/permissions\n2. Check authorization policies\n3. Ensure resource ownership is correctly set\n4. Review permission configuration for the test environment"

        elif '500' in failure_msg or 'internal server error' in failure_msg:
            failure.failure_category = "API Error 500"
            failure.failure_reason = "Server-side error (500). API encountered an unexpected exception."
            failure.solution = "1. Check server logs for stack traces\n2. Verify all required services are running\n3. Check database connectivity\n4. Review API error handling code\n5. File a bug if this is an API defect"

        elif 'timeout' in failure_msg:
            failure.failure_category = "Timeout Failure"
            failure.failure_reason = "Request timeout. API took too long to respond."
            failure.solution = "1. Check API performance and query optimization\n2. Increase timeout threshold if operation is legitimately slow\n3. Verify network connectivity\n4. Check if backend services are under load"

        elif 'exists' in failure_msg and 'not' in failure_msg:
            failure.failure_category = "Missing Field Failure"
            failure.failure_reason = "Expected field or property is missing from the response."
            failure.solution = "1. Check if API response structure changed\n2. Verify API version compatibility\n3. Review API documentation for schema changes\n4. Update test to handle new response format"

        elif '400' in failure_msg or 'bad request' in failure_msg:
            failure.failure_category = "API Error 400"
            failure.failure_reason = "Bad request (400). API rejected the request due to invalid parameters or payload."
            failure.solution = "1. Verify request parameters and body format\n2. Check API documentation for required fields\n3. Validate input data types and constraints\n4. Review API contract for breaking changes"

        elif '502' in failure_msg or 'bad gateway' in failure_msg:
            failure.failure_category = "API Error 502"
            failure.failure_reason = "Bad Gateway (502). Upstream service returned invalid response."
            failure.solution = "1. Check upstream service health\n2. Verify service configuration and routing\n3. Review load balancer and proxy settings\n4. Check service dependencies availability"

        elif 'null' in failure_msg or 'undefined' in failure_msg:
            failure.failure_category = "Null/Undefined Value Failure"
            failure.failure_reason = "Expected value is null or undefined when it should have a value."
            failure.solution = "1. Check data initialization in test setup\n2. Verify API response includes required fields\n3. Review database records for missing data\n4. Check for timing issues in data availability"

        elif 'connection' in failure_msg or 'network' in failure_msg:
            failure.failure_category = "Network/Connection Failure"
            failure.failure_reason = "Network or connection issue prevented request completion."
            failure.solution = "1. Verify service is running and accessible\n2. Check network connectivity\n3. Review firewall and security settings\n4. Validate service URL and port configuration"

        elif 'schema' in failure_msg or 'validation' in failure_msg:
            failure.failure_category = "Schema Validation Failure"
            failure.failure_reason = "Response does not match expected schema or validation rules."
            failure.solution = "1. Review API response schema\n2. Check for API contract changes\n3. Verify schema definition is up to date\n4. Compare expected vs actual response structure"

        else:
            # More specific generic categorization based on context
            if any(status in failure_msg for status in ['200', '201', '204']):
                failure.failure_category = "Assertion Failure (Success Response)"
            elif any(word in failure_msg for word in ['expected', 'actual', 'should', 'equal', 'match']):
                failure.failure_category = "Assertion Mismatch"
            else:
                failure.failure_category = "General Test Failure"

            failure.failure_reason = "Assertion failure. Test expectation did not match actual API behavior."
            failure.solution = "1. Review the full test logs for context\n2. Manually test the API endpoint\n3. Compare expected vs actual response\n4. Determine if test needs updating or API has a bug"

    def analyze_test(self, test_dir: Path) -> Optional[TestFailure]:
        """Analyze a single test directory for failures"""
        test_js = test_dir / 'test.js'
        if not test_js.exists():
            return None

        test_data = self.parse_test_js(test_js)
        if not test_data:
            return None

        # Count this test
        self.total_tests += 1

        # Find failures
        elements = test_data.get('reportElements') or []
        failed_elements = [(i, e) for i, e in enumerate(elements)
                          if e.get('status') == 'failure']

        if not failed_elements:
            self.total_passed += 1
            return None

        self.total_failed += 1

        # Process first failure
        failed_idx, failed_elem = failed_elements[0]

        # Extract test name from test data (try multiple fields)
        test_name = test_data.get('title') or test_data.get('name') or test_data.get('uid') or test_dir.name

        # Create failure object
        failure = TestFailure(test_name, str(test_dir))
        failure.failure_message = failed_elem.get('title', 'Unknown failure')

        # Extract API details
        api_call, api_url, api_method, request_params, response_headers, response_body = \
            self.extract_api_details(test_data, failed_idx, test_dir)

        failure.api_call = api_call
        failure.api_url = api_url
        failure.api_method = api_method
        failure.request_params = request_params
        failure.response_headers = response_headers
        failure.response_body = response_body

        # Perform RCA (pass test_data for 503 detection)
        self.analyze_failure(failure, test_data)

        return failure

    def analyze_all_tests(self):
        """Analyze all tests in the report"""
        test_dirs = self.find_test_directories()

        print(f"\nAnalyzing {len(test_dirs)} tests...")
        for test_dir in test_dirs:
            failure = self.analyze_test(test_dir)
            if failure:
                self.failures.append(failure)

        print(f"\nAnalysis complete:")
        print(f"  Total Tests: {self.total_tests}")
        print(f"  Passed: {self.total_passed}")
        print(f"  Failed: {self.total_failed}")
        print(f"  Analyzed (including 503 errors): {len(self.failures)}")

    def format_api_request_response(self, failure: TestFailure) -> str:
        """Format the API request/response for Excel cell"""
        output = []

        # Request section
        output.append("REQUEST:")
        output.append(f"Method: {failure.api_method}")
        output.append(f"URL: {failure.api_url}")
        if failure.request_params:
            output.append(f"Parameters: {failure.request_params}")
        output.append("")

        # Response section
        output.append("RESPONSE:")
        if failure.response_headers:
            output.append("Response Headers:")
            for key, value in failure.response_headers.items():
                output.append(f"  {key}: {value}")

        if failure.response_body:
            output.append("Body:")
            # Indent response body
            for line in failure.response_body.split('\n'):
                output.append(f"  {line}")

        return '\n'.join(output)

    def generate_excel_report(self) -> str:
        """Generate the Excel report with two sheets"""
        wb = Workbook()

        # Sheet 1: Failed Tests Analysis
        ws1 = wb.active
        ws1.title = "Failed Tests Analysis"

        # Define headers
        headers = ['S.No.', 'Test_Name', 'API_Request_Response', 'Failure', 'Failure_Category', 'Failure_Reason', 'Solution']
        ws1.append(headers)

        # Style headers
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add failure data
        for idx, failure in enumerate(self.failures, 1):
            api_req_resp = self.format_api_request_response(failure)

            row = [
                idx,
                failure.test_name,
                api_req_resp,
                failure.failure_message,
                failure.failure_category,
                failure.failure_reason,
                failure.solution
            ]
            ws1.append(row)

            # Style the row
            row_num = idx + 1
            for col_num in range(1, 8):
                cell = ws1.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Adjust column widths
        ws1.column_dimensions['A'].width = 8
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 50
        ws1.column_dimensions['D'].width = 40
        ws1.column_dimensions['E'].width = 25
        ws1.column_dimensions['F'].width = 40
        ws1.column_dimensions['G'].width = 45

        # Sheet 2: Summary Statistics
        ws2 = wb.create_sheet("Summary Statistics")

        # Calculate stats
        pass_rate = (self.total_passed / self.total_tests * 100) if self.total_tests > 0 else 0

        # Statistics data
        stats = [
            ['Metric', 'Value'],
            ['Total Tests', self.total_tests],
            ['Total Passed', self.total_passed],
            ['Total Failed', self.total_failed],
            ['Tests Analyzed', len(self.failures)],
            ['Pass Rate %', f'{pass_rate:.2f}%'],
            [''],
            ['Failure Categories', 'Count'],
        ]

        # Categorize failures using the failure_category assigned during analysis
        categories = {}
        for failure in self.failures:
            cat = failure.failure_category or 'Uncategorized'
            categories[cat] = categories.get(cat, 0) + 1

        for cat, count in sorted(categories.items()):
            stats.append([cat, count])

        # Write stats to sheet
        for row in stats:
            ws2.append(row)

        # Style statistics sheet
        for row in ws2.iter_rows(min_row=1, max_row=len(stats)):
            for cell in row:
                if cell.row == 1 or cell.row == 9:  # Headers
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 15

        # Save workbook
        output_file = Path(self.output_dir) / f"Test_Failure_Analysis_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        wb.save(output_file)

        print(f"\nExcel report generated: {output_file}")
        return str(output_file)


def main():
    """Main entry point"""
    if len(sys.argv) < 2:
        print("Usage: python analyze_report.py <path_to_zip_file> [output_directory]")
        print(f"Default output directory: {Path.home() / '.claude' / 'outputs'}")
        sys.exit(1)

    zip_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(zip_path):
        print(f"Error: File not found: {zip_path}")
        sys.exit(1)

    analyzer = ReportAnalyzer(zip_path, output_dir)

    try:
        analyzer.extract_report()
        analyzer.analyze_all_tests()
        excel_file = analyzer.generate_excel_report()

        print(f"\n{'='*60}")
        print("ANALYSIS COMPLETE")
        print(f"{'='*60}")
        print(f"Report saved to: {excel_file}")
        print(f"\nSummary:")
        print(f"  - {len(analyzer.failures)} failures analyzed and documented (including 503 errors)")
        print(f"  - Pass rate: {(analyzer.total_passed/analyzer.total_tests*100):.1f}%")

    except Exception as e:
        print(f"\nError during analysis: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
